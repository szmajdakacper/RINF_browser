"""Generator formularza zgodnosci pojazdu z linia.

Refaktor logiki z `fill_form.py` (na pulpicie) tak, aby uzyc sparsowanego
RINFDataset zamiast bezposrednio XML.
"""
from __future__ import annotations
from collections import defaultdict
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import openpyxl

from .. import FOOTER_TEXT
from ..models import RINFDataset, SectionOfLine, OperationalPoint, Parameter
from ..formatters import merge_param_values

# Domyslna sciezka do szablonu: szukaj obok aplikacji, potem na pulpicie
_HERE = Path(__file__).resolve().parent.parent.parent
_LOCAL_TEMPLATE = _HERE / "Formularz zgodności.xlsx"
_DESKTOP_TEMPLATE = Path(r"C:\Users\szmaj\Desktop\RINF\Formularz zgodności.xlsx")
TEMPLATE_PATH_DEFAULT = _LOCAL_TEMPLATE if _LOCAL_TEMPLATE.exists() else _DESKTOP_TEMPLATE

# Mapowanie wierszy formularza (RINF + nazwa) -> lista XML_ID
PARAM_MAP: dict[tuple[str, str], list[str]] = {
    ("1.2.1.0.6.4.", "Dlugosc uzytkowa peronu"):                 ["IPL_Length"],
    ("1.1.1.3.2.9.", "Zgodnosc systemu ETCS"):                    ["CPE_SystemCompatibility"],
    ("1.1.1.3.2.8",  "Pokladowe potwierdzenie ciaglosci pociagu"):["CPE_OptionalFunctions"],
    ("1.1.1.3.2.10.","ETCS M_version"):                           ["CPE_MVersion"],
    ("1.1.1.3.2.11.","Informacje o bezpiecznej dlugosci skladu"): ["CPE_OptionalFunctions", "CPE_RestrictionsConditions"],
    ("1.1.1.3.3.9",  "Zgodnosc lacznosci radiowej - glos"):       ["CRG_RadioCompVoice"],
    ("1.1.1.3.3.10", "Zgodnosc lacznosci radiowej - dane"):       ["CRG_RadioCompData"],
    ("1.1.1.3.3.11.","Sieci GSM-R - umowa roamingowa"):           ["CRG_RoamingAgreement", "CRG_RoamingPublic"],
    ("1.1.1.3.3.5.", "Sieci GSM-R"):                              ["CRG_RoamingAgreement", "CRG_RoamingPublic"],
    ("1.1.1.3.3.4.", "Wykorzystanie grupy 555 przez GSM-R"):      ["CRG_Needof555", "CRG_OptionalFunctions"],
    ("1.1.1.1.3.6 (promil)",  "Promil nachylenia"):               ["ILL_Gradient"],
    ("1.1.1.1.6.1.", "Maksymalne opoznienie pociagu"):            ["ILR_MaxDeceleration"],
    ("1.1.1.3.11.1", "Wymagana maksymalna droga hamowania"):      ["CBP_MaxBrakeDist"],
    ("1.1.1.1.3.6 (profil)",  "Profil nachylenia"):               ["ILL_GradProfile"],
    ("1.1.1.2.2.4.", "Pozwolenie na hamowanie odzyskowe"):        ["ECS_RegenerativeBraking"],
    ("1.1.1.3.5.3.", "Dotychczasowe systemy kontroli pociagu"):   ["CPO_LegacyTrainProtection"],
    ("1.1.1.3.6.1.", "Dotychczasowy system lacznosci radiowej"):  ["CRS_Installed"],
    ("1.1.1.1.3.7",  "Minimalny promien luku poziomego"):         ["ILL_MinRadHorzCurve"],
    ("1.1.1.2.2.1.1.","Rodzaj sieci trakcyjnej"):                 ["ECS_SystemType"],
    ("1.1.1.2.2.1.2.","System zasilania"):                        ["ECS_VoltFreq"],
    ("1.1.1.1.2.4",  "Pojemnosc ladunkowa"):                      ["IPP_LoadCap"],
    ("1.1.1.2.5.1.", "Ograniczenie poboru pradu / mocy"):         ["ERS_PowerLimitOnBoard"],
    ("1.1.1.2.2.5",  "Maksymalna wysokosc przewodu jezdnego"):    ["ECS_MaxWireHeight"],
    ("1.1.1.2.2.6",  "Minimalna wysokosc przewodu jezdnego"):     ["ECS_MinWireHeight"],
    ("1.1.1.2.3.1",  "Dopuszczone slizgacze TSI"):                ["EPA_TSIHeads"],
    ("1.1.1.2.3.2",  "Dopuszczone inne slizgacze"):               ["EPA_OtherHeads"],
    ("1.1.1.2.3.3",  "Liczba uniesionych pantografow / rozstaw"): ["EPA_NumRaisedSpeed"],
    ("1.1.1.2.3.4",  "Material nakladki stykowej"):               ["EPA_StripMaterial"],
    ("1.1.1.2.5.2",  "Dozwolona sila nacisku"):                   ["ERS_ContactForce"],
    ("1.1.1.2.5.3.", "Automatyczna regulacja wysokosci pantografu"):["ERS_AutoDropRequired"],
    ("1.1.1.2.2.3",  "Max pobor pradu na pantograf - postoj"):    ["ECS_MaxStandstillCurrent"],
    ("1.1.1.1.3.1.1.","Skrajnia"):                                ["ILL_Gauging"],
    ("1.1.1.1.7.1",  "Smarowanie obrzezy kol"):                   ["IHS_FlangeLubeForbidden", "CTD_FlangeLubeRules"],
    ("1.1.1.3.7.1.1.","Rodzaj systemu detekcji pociagu"):         ["CTD_DetectionSystem"],
    ("1.1.1.1.6.3.", "Hamulce magnetyczne"):                      ["ILR_MagneticBrakes"],
    ("1.1.1.1.6.2.", "Hamulce wiropradowe"):                      ["ILR_EddyCurrentBrakes"],
    ("1.1.1.1.8.10.","Kategoria pozarowa taboru (tunele)"):       ["ITU_FireCatReq"],
    ("1.1.1.1.8.11.","Krajowa kategoria pozarowa taboru"):        ["ITU_NatFireCatReq"],
    ("1.1.1.1.2.6",  "Zakres temperatur"):                        ["IPP_TempRange"],
    ("1.1.1.1.2.8",  "Wystapienie trudnych warunkow klimatycznych"):["IPP_SevereClimateCon"],
    ("1.1.1.1.2.5",  "Predkosc maksymalna"):                      ["IPP_MaxSpeed"],
    ("1.1.1.1.4.2.", "Niedobor przechylki"):                      ["ITP_CantDeficiency"],
    ("1.1.1.1.4.3.", "Pochylenie poprzeczne szyny"):              ["ITP_RailInclination"],
    ("1.1.1.1.7.4",  "HABD - przytorowe czujniki zagrzanych lozysk"):["IHS_HABDExist"],
    ("1.2.1.0.6.5.", "Wysokosc peronu"):                          ["IPL_Height"],
    ("1.1.1.1.5.2",  "Minimalna srednica kola (krzyzownice)"):    ["ISC_MinWheelDiaFixObtuseCrossings"],
    ("1.1.1.1.4.1.", "Nominalna szerokosc toru"):                 ["ITP_NomGauge"],
    ("1.2.0.0.0.5",  "Lokalizacja geograficzna OP"):              ["__OPGeoLocation__"],
    ("1.2.0.0.0.4.1.","Rodzaje systemow zmiany szerokosci toru"): [],
}


def find_route(ds: RINFDataset, line_id: str, op_start_id: str, op_end_id: str) -> list[SectionOfLine]:
    """BFS po SOL-ach na wskazanej linii pomiedzy dwoma OP.

    Zwraca uporzadkowana liste SOL, ktore prowadza od op_start_id do op_end_id.
    """
    sols = ds.sol_by_line.get(line_id, [])
    if not sols:
        return []

    adj: dict[str, list[tuple[str, SectionOfLine]]] = defaultdict(list)
    for s in sols:
        adj[s.op_start_id].append((s.op_end_id, s))
        adj[s.op_end_id].append((s.op_start_id, s))

    # BFS
    from collections import deque
    prev: dict[str, tuple[str, SectionOfLine]] = {}
    visited = {op_start_id}
    q = deque([op_start_id])
    found = False
    while q:
        cur = q.popleft()
        if cur == op_end_id:
            found = True
            break
        for nb, sol in adj.get(cur, []):
            if nb in visited:
                continue
            visited.add(nb)
            prev[nb] = (cur, sol)
            q.append(nb)
    if not found:
        return []
    # rekonstrukcja
    chain: list[SectionOfLine] = []
    node = op_end_id
    while node != op_start_id:
        from_node, sol = prev[node]
        chain.append(sol)
        node = from_node
    chain.reverse()
    return chain


def collect_values_per_track(ds: RINFDataset, chain: list[SectionOfLine]) -> dict[str, dict[str, set[str]]]:
    """Dla kazdego toru (N/P) zwraca slownik ID -> zbior wartosci display."""
    per_track: dict[str, dict[str, set[str]]] = {"N": defaultdict(set), "P": defaultdict(set)}
    shared: dict[str, set[str]] = defaultdict(set)

    # Geolokacja + perony + tunele OP zbieramy od start->end OP-ow
    op_ids_in_route: set[str] = set()
    for s in chain:
        op_ids_in_route.add(s.op_start_id)
        op_ids_in_route.add(s.op_end_id)
        # SOL-tracks
        for t in s.tracks:
            tid = t.identification
            if tid in per_track:
                for p in t.parameters:
                    per_track[tid][p.id].add(p.display)
            # SOL tunele - wspolne
            for tu in t.tunnels:
                for p in tu.parameters:
                    shared[p.id].add(p.display)

    for oid in op_ids_in_route:
        op = ds.op_by_id.get(oid)
        if not op:
            continue
        if op.geographic and op.geographic.latitude and op.geographic.longitude:
            shared["__OPGeoLocation__"].add(f"{op.name}: {op.geographic.latitude}, {op.geographic.longitude}")
        for t in op.tracks:
            for pl in t.platforms:
                for p in pl.parameters:
                    shared[p.id].add(p.display)
            for tu in t.tunnels:
                for p in tu.parameters:
                    shared[p.id].add(p.display)

    # Polacz shared do obu torow
    for tid in ("N", "P"):
        for pid, vals in shared.items():
            per_track[tid][pid] = per_track[tid].get(pid, set()) | vals
    return per_track


def fill_workbook(template_path: Path, track_id: str, track_vals: dict[str, set[str]], header: dict[str, str]) -> bytes:
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Arkusz2"]

    ws["B3"] = header.get("trasa", "")
    # Komorka D3 ma "Tor: parzysty / nieparzysty *"
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if isinstance(v, str) and "parzysty" in v and "nieparzysty" in v:
            ws.cell(row=3, column=c).value = header.get("tor_label", track_id)
            break
    ws["B4"] = header.get("zarzadca", "PKP PLK S.A.")
    ws["B5"] = header.get("linia", "")
    ws["B6"] = header.get("opracowal", "RINF Browser")
    ws["B7"] = header.get("data", date.today().isoformat())

    for row in range(11, ws.max_row + 1):
        rinf_num = ws.cell(row=row, column=2).value
        name = ws.cell(row=row, column=3).value
        if not rinf_num:
            continue
        rinf_num_str = str(rinf_num).strip()
        name_str = str(name).strip() if name else ""

        matched_ids: list[str] = []
        for (key_num, _key_name), ids in PARAM_MAP.items():
            if key_num.startswith(rinf_num_str.rstrip(".")) or rinf_num_str.startswith(key_num.rstrip(".")):
                if key_num.startswith("1.1.1.1.3.6"):
                    if "profil" in name_str.lower() and "profil" not in key_num.lower():
                        continue
                    if "promil" in name_str.lower() and "promil" not in key_num.lower():
                        continue
                matched_ids.extend(ids)

        merged: set[str] = set()
        for xid in matched_ids:
            if xid in track_vals:
                merged |= track_vals[xid]
        ws.cell(row=row, column=4).value = merge_param_values(merged)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_route_form(
    ds: RINFDataset,
    line_id: str,
    op_start_id: str,
    op_end_id: str,
    template_path: Path = TEMPLATE_PATH_DEFAULT,
) -> dict[str, tuple[bytes, str]]:
    """Generuje dwa pliki xlsx (tor N i P). Zwraca {tor: (bytes, filename)}."""
    chain = find_route(ds, line_id, op_start_id, op_end_id)
    if not chain:
        return {}
    per_track = collect_values_per_track(ds, chain)
    start_op = ds.op_by_id.get(op_start_id)
    end_op = ds.op_by_id.get(op_end_id)
    route_label = f"{start_op.name if start_op else op_start_id} - {end_op.name if end_op else op_end_id}"

    line_num = ""
    if chain:
        line_num = chain[0].line_number_pl

    results: dict[str, tuple[bytes, str]] = {}
    for tid, label in (("N", "nieparzysty"), ("P", "parzysty")):
        header = {
            "trasa": route_label,
            "tor_label": f"Tor: {label} ({tid})",
            "zarzadca": "PKP PLK S.A.",
            "linia": f"{line_num}   |   Tor {label} ({tid})",
            "opracowal": FOOTER_TEXT,
        }
        data = fill_workbook(template_path, tid, per_track[tid], header)
        fname = f"Formularz zgodnosci - linia {line_num} {route_label} - tor {label} ({tid}).xlsx"
        results[tid] = (data, fname)
    return results
